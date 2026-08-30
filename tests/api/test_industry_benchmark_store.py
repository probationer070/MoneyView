import openpyxl
import pytest

from apps.api.services.industry_benchmark_store import (
    latest_vintage,
    load_vintage,
    parse_workbook,
    store_vintage,
)
from packages.core_finance.industry_benchmark import BENCHMARK_COLUMNS, IndustryRow
from tests.fixtures.industry_rows_technology import TECHNOLOGY_ROWS


def _workbook(tmp_path):
    """Build a miniature source workbook. Constructed rather than shipped as a
    binary so the test needs no network and no checked-in .xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Industry Average Beta (US)"
    headers = ["Industry Name", "Number of firms"] + [c.source_header for c in BENCHMARK_COLUMNS]
    ws.append(headers)
    for row in TECHNOLOGY_ROWS[:3]:
        ws.append([row.name, row.firms] + [row.values[c.key] for c in BENCHMARK_COLUMNS])
    ws.append(["Total Market", 5994] + [0.1] * len(BENCHMARK_COLUMNS))
    path = tmp_path / "industries.xlsx"
    wb.save(path)
    return str(path)


def test_parse_reads_rows_by_header_name_not_position(tmp_path):
    rows = parse_workbook(_workbook(tmp_path))
    by_name = {r.name: r for r in rows}
    assert by_name["Semiconductor Equip"].firms == 31
    assert by_name["Semiconductor Equip"].values["after_tax_roc"] == pytest.approx(0.2840446307)


def test_parse_excludes_aggregate_rows(tmp_path):
    assert "Total Market" not in {r.name for r in parse_workbook(_workbook(tmp_path))}


def test_parse_raises_when_a_required_header_is_missing(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Industry Average Beta (US)"
    wb.active.append(["Industry Name", "Number of firms"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="missing"):
        parse_workbook(str(path))


def test_store_and_load_round_trip():
    assert store_vintage("2026-01-01", TECHNOLOGY_ROWS[:3]) == 3
    loaded = {r.name: r for r in load_vintage("2026-01-01")}
    assert loaded["Computers/Peripherals"].firms == 36
    assert loaded["Computers/Peripherals"].values["reinvestment_rate"] == pytest.approx(0.2136889975)


def test_storing_the_same_vintage_twice_replaces_rather_than_duplicates():
    store_vintage("2026-01-01", TECHNOLOGY_ROWS[:3])
    store_vintage("2026-01-01", TECHNOLOGY_ROWS[:3])
    assert len(load_vintage("2026-01-01")) == 3


def test_latest_vintage_returns_the_newest_at_or_before_a_date():
    store_vintage("2025-01-01", TECHNOLOGY_ROWS[:3])
    store_vintage("2026-01-01", TECHNOLOGY_ROWS[:3])
    assert latest_vintage() == "2026-01-01"
    assert latest_vintage(on_or_before="2025-06-01") == "2025-01-01"
    assert latest_vintage(on_or_before="2024-01-01") is None


def test_latest_vintage_is_none_when_nothing_was_ever_stored():
    assert latest_vintage() is None


def test_store_and_load_round_trip_real_values_for_the_four_new_columns():
    """test_store_and_load_round_trip above uses TECHNOLOGY_ROWS, whose four new
    keys are all None (that fixture predates these columns), so it never proves a
    real value survives the INSERT/SELECT round trip -- a misaligned placeholder
    or a column dropped from the SELECT would pass it silently."""
    row = IndustryRow("Priced Industry", 20, {
        "revenue_growth": 0.1, "operating_margin": 0.1, "after_tax_roc": 0.1,
        "effective_tax_rate": 0.1, "unlevered_beta": 1.0, "debt_to_capital": 0.1,
        "cost_of_capital": 0.1, "sales_to_capital": 1.0, "reinvestment_rate": 0.1,
        "trailing_pe": 22.5, "price_to_book": 3.5, "ev_sales": 4.5, "stdev_price": 0.3,
    })
    store_vintage("2026-02-01", [row])
    loaded = load_vintage("2026-02-01")[0]
    assert loaded.values["trailing_pe"] == pytest.approx(22.5)
    assert loaded.values["price_to_book"] == pytest.approx(3.5)
    assert loaded.values["ev_sales"] == pytest.approx(4.5)
    assert loaded.values["stdev_price"] == pytest.approx(0.3)


@pytest.mark.virgin_db
def test_init_db_adds_the_four_price_columns_to_a_legacy_industry_benchmark_table(tmp_path, monkeypatch):
    """_CREATE_SCHEMA_SQL already creates industry_benchmark with all thirteen
    columns, so every ordinary test database is born migrated and the guarded
    ALTER TABLE branch in _ensure_schema_compatibility runs on nothing. Build the
    pre-Task-3, nine-column table by hand to actually exercise that migration,
    the same way test_init_db_adds_comparison_universe_columns_for_legacy_snapshot_tables
    (tests/api/test_corporate_comparison.py) exercises its own ALTER TABLE."""
    import sqlite3

    from apps.api.services import db as db_service

    db_path = tmp_path / "moneyview.db"
    monkeypatch.setattr(db_service, "_DB_PATH", db_path)

    with sqlite3.connect(str(db_path)) as conn:
        conn.executescript(
            """
            CREATE TABLE industry_benchmark (
                vintage            TEXT NOT NULL,
                industry_name      TEXT NOT NULL,
                firms              INTEGER NOT NULL,
                revenue_growth     REAL,
                operating_margin   REAL,
                after_tax_roc      REAL,
                effective_tax_rate REAL,
                unlevered_beta     REAL,
                debt_to_capital    REAL,
                cost_of_capital    REAL,
                sales_to_capital   REAL,
                reinvestment_rate  REAL,
                PRIMARY KEY (vintage, industry_name)
            );
            """
        )
        conn.execute(
            """INSERT INTO industry_benchmark
                   (vintage, industry_name, firms, revenue_growth)
               VALUES ('2025-01-01', 'Legacy Industry', 50, 0.05)"""
        )

    db_service.init_db()

    with db_service.get_db() as conn:
        columns = {row["name"] for row in conn.execute("PRAGMA table_info(industry_benchmark)")}
        legacy_row = conn.execute(
            "SELECT * FROM industry_benchmark WHERE industry_name = 'Legacy Industry'"
        ).fetchone()

    assert {"trailing_pe", "price_to_book", "ev_sales", "stdev_price"}.issubset(columns)
    assert legacy_row["revenue_growth"] == pytest.approx(0.05)
    assert legacy_row["trailing_pe"] is None

    # A no-op on an already-migrated table: ALTER TABLE ADD COLUMN without the
    # guard would raise "duplicate column name" the second time init_db() runs.
    db_service.init_db()


def test_a_workbook_without_the_optional_columns_still_parses(tmp_path):
    """The four price columns were added after several vintages were published."""
    import openpyxl

    from apps.api.services.industry_benchmark_store import parse_workbook
    from packages.core_finance.industry_benchmark import BENCHMARK_COLUMNS

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Industry Averages"
    required = [c for c in BENCHMARK_COLUMNS if c.required]
    sheet.append(["Industry Name", "Number of firms"] + [c.source_header for c in required])
    sheet.append(["Semiconductor", 50] + [0.1] * len(required))
    path = tmp_path / "old_vintage.xlsx"
    book.save(path)

    rows = parse_workbook(path, sheet="Industry Averages")
    assert len(rows) == 1
    assert rows[0].values.get("trailing_pe") is None
